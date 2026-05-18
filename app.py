import os
import calendar
from collections import OrderedDict
from datetime import datetime, date, timedelta
from flask import Flask, render_template, request, jsonify, redirect, url_for, make_response, g, send_file
import io
from openpyxl import load_workbook, Workbook
from translations import TRANSLATIONS
from holidays import get_bg_holidays_range, holiday_name, is_substitute

app = Flask(__name__)

_data_dir = os.environ.get("DATA_DIR", os.path.dirname(__file__))
EXCEL_FILE = os.path.join(_data_dir, "attendance.xlsx")


# ── Language ──────────────────────────────────────────────────────────────────

@app.before_request
def set_language():
    g.lang = request.cookies.get("lang", "en")
    if g.lang not in TRANSLATIONS:
        g.lang = "en"
    g.t = TRANSLATIONS[g.lang]


@app.context_processor
def inject_t():
    return {"t": g.t, "lang": g.lang}


@app.route("/set-lang/<lang>")
def set_lang(lang):
    if lang not in TRANSLATIONS:
        lang = "en"
    resp = make_response(redirect(request.referrer or "/"))
    resp.set_cookie("lang", lang, max_age=60 * 60 * 24 * 365)
    return resp


# ── Excel helpers ─────────────────────────────────────────────────────────────

def get_workbook():
    if os.path.exists(EXCEL_FILE):
        return load_workbook(EXCEL_FILE)
    wb = Workbook()
    ws_emp = wb.active
    ws_emp.title = "Сотрудники"
    ws_emp.append(["Имя", "Должность", "Юнит", "Email"])
    ws_att = wb.create_sheet("Посещаемость")
    ws_att.append(["Дата", "Сотрудник", "Статус", "Время"])
    wb.save(EXCEL_FILE)
    return wb


def save_workbook(wb):
    wb.save(EXCEL_FILE)


def get_employees():
    wb = get_workbook()
    ws = wb["Сотрудники"]
    return [
        {"name": str(r[0]), "position": str(r[1]) if r[1] else "", "unit": str(r[2]) if r[2] else ""}
        for r in ws.iter_rows(min_row=2, values_only=True) if r[0]
    ]


def get_employees_by_unit():
    groups = OrderedDict()
    for emp in get_employees():
        groups.setdefault(emp["unit"] or "—", []).append(emp)
    return groups


def get_attendance(year, month):
    wb = get_workbook()
    ws = wb["Посещаемость"]
    prefix = f"{year}-{month:02d}"
    return [
        {"Дата": str(r[0]), "Сотрудник": str(r[1]) if r[1] else "",
         "Статус": str(r[2]) if r[2] else "", "Время": str(r[3]) if r[3] else ""}
        for r in ws.iter_rows(min_row=2, values_only=True)
        if r[0] and str(r[0]).startswith(prefix)
    ]


def save_attendance(employee, status):
    wb = get_workbook()
    ws = wb["Посещаемость"]
    today = date.today().isoformat()
    now_time = datetime.now().strftime("%H:%M")
    for row in ws.iter_rows(min_row=2):
        if str(row[0].value) == today and str(row[1].value) == employee:
            row[2].value = status
            row[3].value = now_time
            save_workbook(wb)
            return
    ws.append([today, employee, status, now_time])
    save_workbook(wb)


# ── Calendar helpers ──────────────────────────────────────────────────────────

def build_calendar_cells(year, month, employee_names, attendance_map, holidays):
    today_iso = date.today().isoformat()
    cells = []
    for week in calendar.monthcalendar(year, month):
        for weekday, day in enumerate(week):
            if day == 0:
                cells.append({"day": 0})
            else:
                day_iso = f"{year}-{month:02d}-{day:02d}"
                cells.append({
                    "day": day, "weekday": weekday,
                    "is_today": day_iso == today_iso,
                    "holiday": holiday_name(holidays, day_iso),
                    "is_substitute": is_substitute(holidays, day_iso),
                    "statuses": [
                        (n, attendance_map[(day_iso, n)])
                        for n in employee_names if (day_iso, n) in attendance_map
                    ],
                })
    return cells


def build_week_days():
    today = date.today()
    monday = today - timedelta(days=today.weekday())
    return [
        {
            "iso": (monday + timedelta(days=i)).isoformat(),
            "idx": i,
            "date": (monday + timedelta(days=i)).strftime("%d.%m"),
            "is_today": (monday + timedelta(days=i)) == today,
        }
        for i in range(7)
    ]


def month_cells(y, m, planned, today_iso, holidays):
    cells = []
    for week in calendar.monthcalendar(y, m):
        for weekday, day in enumerate(week):
            if day == 0:
                cells.append(None)
            else:
                d_iso = f"{y}-{m:02d}-{day:02d}"
                cells.append({
                    "day": day, "iso": d_iso, "weekday": weekday,
                    "is_past": d_iso < today_iso,
                    "is_today": d_iso == today_iso,
                    "status": planned.get(d_iso, ""),
                    "holiday": holiday_name(holidays, d_iso),
                    "is_substitute": is_substitute(holidays, d_iso),
                })
    return cells


# ── Routes ────────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html",
                           employees_by_unit=get_employees_by_unit(),
                           today=date.today().isoformat())


@app.route("/checkin", methods=["POST"])
def checkin():
    employee = request.form.get("employee", "").strip()
    status = request.form.get("status", "").strip()
    if not employee or not status:
        return jsonify({"error": g.t["checkin_pick_status_err"]}), 400
    save_attendance(employee, status)
    return jsonify({"ok": True, "message": f"{employee} — {status}"})


@app.route("/dashboard")
def dashboard():
    employees = get_employees()
    names = [e["name"] for e in employees]
    today = date.today()
    year = int(request.args.get("year", today.year))
    month = int(request.args.get("month", today.month))
    view = request.args.get("view", "month")

    records = get_attendance(year, month)
    attendance_map = {(r["Дата"], r["Сотрудник"]): r["Статус"] for r in records}
    today_iso = today.isoformat()
    holidays = get_bg_holidays_range(year, year)

    return render_template("dashboard.html",
        employees=names,
        attendance_map=attendance_map,
        year=year, month=month,
        month_name=g.t["months"][month],
        view=view,
        today=today_iso,
        today_office=sum(1 for n in names if attendance_map.get((today_iso, n)) == "Офис"),
        today_remote=sum(1 for n in names if attendance_map.get((today_iso, n)) == "Удалёнка"),
        calendar_cells=build_calendar_cells(year, month, names, attendance_map, holidays),
        week_days=build_week_days(),
        holidays=holidays,
    )


@app.route("/plan")
def plan():
    today = date.today()
    year = int(request.args.get("year", today.year))
    month = int(request.args.get("month", today.month))
    employee_name = request.args.get("employee", "")

    next_month = month + 1 if month < 12 else 1
    next_year = year if month < 12 else year + 1

    planned = {}
    if employee_name:
        for r in get_attendance(year, month) + get_attendance(next_year, next_month):
            if r["Сотрудник"] == employee_name:
                planned[r["Дата"]] = r["Статус"]

    today_iso = today.isoformat()
    t = g.t
    holidays = get_bg_holidays_range(
        min(year, next_year), max(year, next_year)
    )

    return render_template("plan.html",
        employees_by_unit=get_employees_by_unit(),
        employee_name=employee_name,
        year=year, month=month,
        next_year=next_year, next_month=next_month,
        month_name=t["months"][month],
        next_month_name=t["months"][next_month],
        cells=month_cells(year, month, planned, today_iso, holidays),
        next_cells=month_cells(next_year, next_month, planned, today_iso, holidays),
        today=today_iso,
        holidays=holidays,
    )


@app.route("/plan/save", methods=["POST"])
def plan_save():
    data = request.get_json()
    employee = (data.get("employee") or "").strip()
    days = data.get("days", {})
    if not employee:
        return jsonify({"error": "No employee"}), 400

    wb = get_workbook()
    ws = wb["Посещаемость"]
    existing = {}
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        existing[(str(row[0].value), str(row[1].value))] = i

    for d_iso, status in days.items():
        key = (d_iso, employee)
        if status == "":
            if key in existing:
                ws.delete_rows(existing[key])
                existing = {k: v for k, v in existing.items() if v != existing[key]}
        elif key in existing:
            idx = existing[key]
            ws.cell(idx, 3).value = status
            ws.cell(idx, 4).value = "план"
        else:
            ws.append([d_iso, employee, status, "план"])

    save_workbook(wb)
    return jsonify({"ok": True, "saved": len(days)})


@app.route("/export")
def export():
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    today = date.today()
    year  = int(request.args.get("year",  today.year))
    month = int(request.args.get("month", today.month))

    employees  = get_employees()
    emp_names  = [e["name"] for e in employees]
    records    = get_attendance(year, month)
    att_map    = {(r["Дата"], r["Сотрудник"]): r["Статус"] for r in records}
    holidays   = get_bg_holidays_range(year, year)

    import calendar as cal_mod
    _, n_days = cal_mod.monthrange(year, month)
    days = [date(year, month, d) for d in range(1, n_days + 1)]

    # ── Styles ────────────────────────────────────────────────────────────────
    ORANGE   = "FFFF6600"
    NAVY     = "FF1D1F2E"
    OFFICE_C = "FFD4F5DF"
    REMOTE_C = "FFFFF3D6"
    ABSENT_C = "FFFFD6D4"
    HOLIDAY_C= "FFFFF8F0"
    WEEKEND_C= "FFF5F5F7"
    GREY_C   = "FFF5F5F7"

    thin = Side(style="thin", color="FFE8E8EC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def hdr_font(bold=True, color="FFFFFFFF", size=10):
        return Font(name="Inter", bold=bold, color=color, size=size)

    def cell_font(bold=False, color="FF1D1F2E", size=9):
        return Font(name="Inter", bold=bold, color=color, size=size)

    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center")

    # ── Workbook ──────────────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    t  = g.t
    ws.title = f"{t['months'][month]} {year}"

    month_name_str = t["months"][month]
    STATUS = {
        "Офис":       (t["status_office"],  OFFICE_C),
        "Удалёнка":   (t["status_remote"],  REMOTE_C),
        "Отсутствую": (t["status_absent"],  ABSENT_C),
    }

    # ── Row 1: title ──────────────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1,   end_column=2 + len(days))
    title_cell = ws.cell(1, 1, f"{month_name_str} {year}")
    title_cell.fill      = fill(NAVY)
    title_cell.font      = Font(name="Inter", bold=True, color="FFFFFFFF", size=13)
    title_cell.alignment = center
    ws.row_dimensions[1].height = 28

    # ── Row 2: column headers ─────────────────────────────────────────────────
    ws.cell(2, 1, "#").fill      = fill(ORANGE)
    ws.cell(2, 1).font           = hdr_font()
    ws.cell(2, 1).alignment      = center
    ws.cell(2, 2, t["nav_employees"]).fill = fill(ORANGE)
    ws.cell(2, 2).font           = hdr_font()
    ws.cell(2, 2).alignment      = left

    day_labels = t["days"]  # ["Пн","Вт",...]
    for col_i, d in enumerate(days, start=3):
        c = ws.cell(2, col_i)
        c.value     = d.day
        c.alignment = center
        iso = d.isoformat()
        if iso in holidays:
            c.fill = fill(HOLIDAY_C)
            c.font = Font(name="Inter", bold=True, color=ORANGE[2:], size=9)
        elif d.weekday() >= 5:
            c.fill = fill(WEEKEND_C)
            c.font = Font(name="Inter", bold=True, color="FFFF3B30", size=9)
        else:
            c.fill = fill(ORANGE)
            c.font = hdr_font()

    # Day-of-week sub-header
    ws.cell(3, 1, "").fill = fill(NAVY)
    ws.cell(3, 2, "").fill = fill(NAVY)
    for col_i, d in enumerate(days, start=3):
        c = ws.cell(3, col_i)
        c.value     = day_labels[d.weekday()]
        c.alignment = center
        iso = d.isoformat()
        if iso in holidays:
            c.fill = fill(HOLIDAY_C)
            c.font = Font(name="Inter", bold=True, color=ORANGE[2:], size=8)
        elif d.weekday() >= 5:
            c.fill = fill(WEEKEND_C)
            c.font = Font(name="Inter", bold=True, color="FFFF3B30", size=8)
        else:
            c.fill = fill(NAVY)
            c.font = hdr_font(size=8)

    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 14

    # ── Data rows ─────────────────────────────────────────────────────────────
    for row_i, emp in enumerate(emp_names, start=4):
        ws.row_dimensions[row_i].height = 18
        # Index
        idx_c = ws.cell(row_i, 1, row_i - 3)
        idx_c.fill      = fill(GREY_C)
        idx_c.font      = cell_font(color="FF6F6F79", size=8)
        idx_c.alignment = center
        idx_c.border    = border
        # Name
        name_c = ws.cell(row_i, 2, emp)
        name_c.font      = cell_font(bold=True)
        name_c.alignment = left
        name_c.border    = border

        for col_i, d in enumerate(days, start=3):
            iso    = d.isoformat()
            status = att_map.get((iso, emp), "")
            c      = ws.cell(row_i, col_i)
            c.alignment = center
            c.border    = border

            if iso in holidays:
                c.fill  = fill(HOLIDAY_C)
                c.value = "🇧🇬"
                c.font  = cell_font(size=9)
            elif d.weekday() >= 5:
                c.fill  = fill(WEEKEND_C)
                c.font  = cell_font(size=9)
            elif status in STATUS:
                label, color = STATUS[status]
                c.fill  = fill(color)
                c.value = label
                c.font  = cell_font(size=8)
            else:
                c.font = cell_font(color="FFD0D0D5", size=8)

    # ── Summary row ───────────────────────────────────────────────────────────
    sum_row = 4 + len(emp_names)
    ws.row_dimensions[sum_row].height = 18
    s_label = ws.cell(sum_row, 2,
        "В офисе" if g.lang == "ru" else
        "В офиса" if g.lang == "bg" else "In office")
    s_label.fill      = fill(NAVY)
    s_label.font      = hdr_font(size=9)
    s_label.alignment = left
    ws.cell(sum_row, 1).fill = fill(NAVY)

    for col_i, d in enumerate(days, start=3):
        iso   = d.isoformat()
        count = sum(1 for n in emp_names if att_map.get((iso, n)) == "Офис")
        c = ws.cell(sum_row, col_i)
        c.border    = border
        c.alignment = center
        if iso in holidays or d.weekday() >= 5:
            c.fill = fill(WEEKEND_C if d.weekday() >= 5 else HOLIDAY_C)
        elif count:
            c.value = count
            c.fill  = fill(OFFICE_C)
            c.font  = cell_font(bold=True, color="FF1A6B30", size=9)
        else:
            c.fill = fill(GREY_C)

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 24
    for col_i in range(3, 3 + len(days)):
        ws.column_dimensions[get_column_letter(col_i)].width = 4.5

    ws.freeze_panes = "C4"

    # ── Stream to browser ─────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"attendance_{year}_{month:02d}.xlsx"
    return send_file(buf, as_attachment=True,
                     download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/setup")
def setup():
    return render_template("setup.html",
                           employees_by_unit=get_employees_by_unit(),
                           excel_path=EXCEL_FILE)


@app.route("/add-employee", methods=["POST"])
def add_employee():
    name = request.form.get("name", "").strip()
    if not name:
        return jsonify({"error": g.t["field_name"]}), 400
    wb = get_workbook()
    wb["Сотрудники"].append([
        name,
        request.form.get("position", "").strip(),
        request.form.get("unit", "").strip(),
        request.form.get("email", "").strip(),
    ])
    save_workbook(wb)
    return jsonify({"ok": True})


@app.route("/delete-employee", methods=["POST"])
def delete_employee():
    name = request.form.get("name", "").strip()
    wb = get_workbook()
    ws = wb["Сотрудники"]
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if str(row[0]) == name:
            ws.delete_rows(i)
            save_workbook(wb)
            return jsonify({"ok": True})
    return jsonify({"error": "Not found"}), 404


if __name__ == "__main__":
    get_workbook()
    port = int(os.environ.get("PORT", 5050))
    debug = os.environ.get("FLASK_ENV") != "production"
    print("\n  Attendance tracker started!")
    print(f"  Data file: {EXCEL_FILE}")
    if debug:
        print(f"  Open in browser: http://localhost:{port}\n")
    app.run(debug=debug, host="0.0.0.0", port=port)
